from collections import defaultdict
from pprint import pprint


import wget
import xlrd
from openpyxl import load_workbook
import scrapy
from tempfile import NamedTemporaryFile

class HessenSpider(scrapy.Spider):
    name = "hessen"

    start_urls = ['https://statistik.hessen.de/sites/statistik.hessen.de/files/Verz-6_19.xlsx']

    def parse(self, response):
        with open('hessen.xlsx', 'wb') as f:
            f.write(response.body)

        return get_hessen_open()

"""
[allgemein_öffentlich]: https://statistik.hessen.de/sites/statistik.hessen.de/files/Verz-6_19.xlsx
[allgemein_privat]: https://statistik.hessen.de/sites/statistik.hessen.de/files/Verz-9_16.xlsx
[beruf_öffentlich]: https://statistik.hessen.de/sites/statistik.hessen.de/files/Verz-7_19.xlsx
[beruf_privat]: https://statistik.hessen.de/sites/statistik.hessen.de/files/Verz-8_16.pdf
"""


def get_hessen_open():
    ###############################################################
    #url_mv = 'https://statistik.hessen.de/sites/statistik.hessen.de/files/Verz-6_19.xlsx'
    #wget.download(url_mv, 'hessen.xlsx')

    wb2 = load_workbook('hessen.xlsx', read_only=True)

    worksheet = wb2['Schulverzeichnis']

    collection = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True)):
        if row_number <= 1:
            continue

        row_data = {
            'Schulnummer': row[0],
            'Kreis': row[1],
            'Gemeinde': row[2],
            'Rechtsform': row[3],
            'Gesamtschule': row[4],
            'Name der Schule': row[5],
            'PLZ': row[6],
            'Schulort': row[7],
            'Adresse': row[8],
            'Telefonvorwahl': row[9],
            'Telefonnummer': row[10],
            'Ausländische Schüler': row[11],
            'Kennung Internat': row[12],
            # 'Vorklassen': row[13],
            # 'Grundschulen': row[14],
            # 'davon an Eingangstufe': row[15],
            # ' Grundschule (inkl. GFLX)': row[16],
            # 'Förderstufe': row[17],
            # 'Hauptschule': row[18],
            # 'Mittelstufen-schule': row[19],
            # 'Realschule': row[20],
            # 'Integrierte Jahrgangs-stufe': row[21],
            # 'Gymnasien': row[22],
            # 'davon an Gymnasien (Mittelstufe)': row[23],
            # ' Gymnasien (Oberstufe)': row[24],
            # 'Förderschulen': row[25],
            # 'Förderschule mit Förderschwerpunkt Sehen': row[26],
            # ' emotionale und soziale Entwicklung': row[27],
            # ' Hören': row[28],
            # ' körperliche und motorische Entwicklung': row[29],
            # ' Kranke': row[30],
            # ' Lernen': row[31],
            # ' Sprachheil-förderung': row[32],
            # ' geistige Entwicklung': row[33],
            # ' Realschul-klassen': row[34],
            # 'Nichtdeutscher Herkunfts-sprache': row[35],
            #            'Schulen für Erwachsene': row[36],
            # 'davon an Abend-hauptschule': row[37],
            # ' Abend-realschule': row[38],
            # ' Abend-gymnasium': row[39],
            # ' Kolleg': row[40],
            # 'Schüler insg. mit Vorklassen': row[41],
            # 'Schüler insg. ohne Vorklassen': row[42],
            'Fax': row[43],
            'Email': row[44],
        }

        yield row_data
        #collection.append(row_data)

    #pprint(collection)


if __name__ == '__main__':
    #get_hessen()
    get_hessen_open()
